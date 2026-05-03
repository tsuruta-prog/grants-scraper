"""Excel 入出力（openpyxl）。

- data/grants.xlsx を読み込み、既存 URL を集合化
- 新規分のみ自治体別シートと統合シートに追記
- 列幅・ヘッダースタイル設定込み
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import ALL_SHEET, COLUMNS, EXCEL_PATH
from utils.logger import get_logger

logger = get_logger(__name__)


HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _ensure_sheet(wb: Workbook, name: str) -> Worksheet:
    """シートが無ければ作成し、ヘッダー行を整える。"""
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    ws.append(COLUMNS)
    for col_idx, _ in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # 列幅
    widths = [12, 40, 14, 14, 18, 60, 50, 20]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"
    return ws


def _load_or_create(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    # デフォルトの "Sheet" を消して統合シートを置く
    default = wb.active
    wb.remove(default)
    _ensure_sheet(wb, ALL_SHEET)
    return wb


def load_existing_urls(path: Path = EXCEL_PATH) -> set[str]:
    """既存ブックから URL の集合を読み出す。新規判定に使う。"""
    if not path.exists():
        return set()
    wb = load_workbook(path, read_only=True)
    urls: set[str] = set()
    try:
        url_idx = COLUMNS.index("URL")  # 0-origin
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) > url_idx and row[url_idx]:
                    urls.add(str(row[url_idx]).strip())
    finally:
        wb.close()
    return urls


def append_grants(grants: Iterable[dict], path: Path = EXCEL_PATH) -> int:
    """grants を自治体別シートと統合シートへ追記。書き込み件数を返す。

    grants の各 dict は config.COLUMNS のキーを持つ前提。
    URL での重複排除は呼び出し側で済ませて渡すこと。
    """
    grants = list(grants)
    if not grants:
        logger.info("追記対象なし")
        return 0

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = _load_or_create(path)

    all_ws = _ensure_sheet(wb, ALL_SHEET)

    written = 0
    for g in grants:
        muni = g.get("自治体名") or "未設定"
        ws = _ensure_sheet(wb, muni)
        row = [g.get(col, "") for col in COLUMNS]
        ws.append(row)
        all_ws.append(row)
        written += 1

    wb.save(path)
    logger.info("Excel 追記 %d 件 -> %s", written, path)
    return written


def now_jst_str() -> str:
    """取得日時として埋める JST 文字列。GitHub Actions では TZ=Asia/Tokyo を設定する想定。"""
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
